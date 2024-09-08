import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def parse_arxml(file_path):
    # Parse the ARXML file
    tree = ET.parse(file_path)
    root = tree.getroot()

    # Initialize a list to store extracted data
    data = []

    # Iterate over elements in the ARXML file (modify this part as per your ARXML structure)
    for element in root.iter():
        tag = element.tag.split('}')[-1]  # Remove namespace
        text = element.text.strip() if element.text else None

        data.append({
            "Tag": tag,
            "Text": text,
        })

    return data


def convert_to_excel(data, output_path):
    df = pd.DataFrame(data)

    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    # Autofit columns based on content
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Autofit rows based on content
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20  # Adjust as needed

    ws.freeze_panes = ws['B2']  # Freezes first row and first column

    # Enable filter option
    ws.auto_filter.ref = ws.dimensions

    fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill

    wb.save(output_path)


# File paths
arxml_file = r"E:\Python_test\arxml_to_xlsx\system.arxml"
output_excel_file = r'E:\Python_test\arxml_to_xlsx\output.xlsx'

# Parse ARXML and convert to Excel with formatting
data = parse_arxml(arxml_file)
convert_to_excel(data, output_excel_file)

print(f"ARXML data has been successfully converted and formatted in {output_excel_file}")
